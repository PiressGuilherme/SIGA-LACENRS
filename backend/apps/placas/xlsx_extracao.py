"""
Geração do mapa de extração preenchendo o template XLSX.

Preenche ``template_mapa_extracao.xlsx`` com dados da placa e retorna os
bytes do XLSX pronto para download — sem conversão para PDF.

Campos preenchidos:
  - D1: código da placa no título (substitui 'xxxxxx')
  - L4: código da placa no ensaio (substitui 'xxxxxx')
  - B4: kit de extração selecionado
  - C6:N13: grid 8×12 com código interno das amostras / CP / CN
  - K17: Registro extração = nome do usuário que gerou o mapa
  - K18: Operador extração = em branco (preenchimento manual)
"""
import io
import os

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock

from .models import TipoConteudoPoco


def _substituir_celula(cell, old, new):
    """Substitui texto numa célula preservando rich text (runs com formatação)."""
    val = cell.value
    if isinstance(val, CellRichText):
        for run in val:
            if isinstance(run, TextBlock) and old in run.text:
                run.text = run.text.replace(old, new)
    elif isinstance(val, str):
        cell.value = val.replace(old, new)

ROWS        = list('ABCDEFGH')
COLS_PADDED = [f'{i:02d}' for i in range(1, 13)]

GRID_START_ROW = 6
GRID_START_COL = 3   # coluna C

TIPO_LABELS = {
    TipoConteudoPoco.CONTROLE_NEGATIVO: 'CN',
    TipoConteudoPoco.CONTROLE_POSITIVO: 'CP',
}

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template_mapa_extracao.xlsx')


def _get_label(poco):
    label = TIPO_LABELS.get(poco.tipo_conteudo)
    if label is not None:
        return label
    if poco.amostra:
        return poco.amostra.codigo_interno
    return ''


def gerar_xlsx_extracao(placa, operador=None, kit_nome=''):
    """Preenche o template e retorna bytes do XLSX.

    Args:
        placa:    instância de Placa (tipo extração)
        operador: instância de User — quem gerou o mapa ("Registro extração")
        kit_nome: nome do kit de extração selecionado
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH, rich_text=True)
    ws = wb.active

    # Código da placa no título e ensaio (preserva rich text / formatação dos runs)
    _substituir_celula(ws['D1'], 'xxxxxx', placa.codigo)
    _substituir_celula(ws['L4'], 'xxxxxx', placa.codigo)

    # Kit de extração
    if kit_nome:
        ws['B4'].value = f'Kit: {kit_nome}'

    # Grid 8×12
    pocos_map = {p.posicao: p for p in placa.pocos.select_related('amostra').all()}
    for ri, row_letter in enumerate(ROWS):
        for ci, col_num in enumerate(COLS_PADDED):
            pos = f'{row_letter}{col_num}'
            poco = pocos_map.get(pos)
            cell = ws.cell(row=GRID_START_ROW + ri, column=GRID_START_COL + ci)
            if poco and poco.tipo_conteudo != TipoConteudoPoco.VAZIO:
                cell.value = _get_label(poco)
            else:
                cell.value = None

    # Registro extração = quem gerou o mapa
    registro_nome = ''
    if operador:
        registro_nome = getattr(operador, 'nome_completo', str(operador))
    ws['K17'].value = f'Registro extração: {registro_nome}' if registro_nome else 'Registro extração:'
    ws['K18'].value = 'Operador extração:'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Mantém compatibilidade com imports antigos
gerar_pdf_extracao = gerar_xlsx_extracao
