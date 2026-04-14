"""
Parser do XLSX de resultados do Sistema de PCR em Tempo Real – Amplio® 96.

Formato do XLSX (aba "0"):
  - Linha de cabeçalho: (None, Well, Fluor, Content, Target, Sample,
                         Threshold Cycle(C(t)), C(t) Mean, C(t) Std. Dev,
                         Starting Quantity (SQ), Log Starting Quantity, SQ Mean, SQ Std. Dev)
  - Múltiplas linhas por poço (uma por canal de fluorescência)
  - Cq = 0.0 significa "sem amplificação" (equivalente ao NaN do CFX)

Mapeamento de Content:
  - Unknown  → amostra
  - Positive → cp
  - Negative → cn

Mapeamento de Target (mesmo que o CFX):
  - HPV-16 / HPV 16 / HPV16 → HPV16
  - HPV-18 / HPV 18 / HPV18 → HPV18
  - HPV AR / HPV-AR / HPV_AR → HPV_AR
  - CI                       → CI

Posição de poço: "A1" → normalizado para "A01" (padrão do sistema).
"""
from typing import Optional

from apps.resultados.parser import TARGET_CANAL_MAP, _CANAIS, resolver_canal

# ── Mapeamento de tipos de conteúdo ──────────────────────────────────────────
CONTENT_TYPE_MAP: dict[str, str] = {
    'Unknown':  'amostra',
    'Positive': 'cp',
    'Negative': 'cn',
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalizar_posicao(well: str) -> str:
    """
    Converte o formato de poço do Amplio® ("A1", "H12") para o padrão
    do sistema ("A01", "H12").
    """
    well = well.strip()
    if len(well) >= 2:
        letra = well[0].upper()
        num   = well[1:]
        try:
            return f'{letra}{int(num):02d}'
        except ValueError:
            pass
    return well


def _safe_cq(val) -> Optional[float]:
    """
    Converte o valor de Ct para float.
    Retorna None quando o valor for 0.0 (sem amplificação no Amplio®) ou inválido.
    """
    if val is None:
        return None
    try:
        f = float(val)
        return None if f == 0.0 else f
    except (TypeError, ValueError):
        return None


# ── Parser principal ──────────────────────────────────────────────────────────

def parse_amplio_xlsx(content: bytes, filename: str = '') -> dict:
    """
    Parseia o XLSX do Amplio® 96.

    Retorna a mesma estrutura que ``parse_cfx_csv``::

        {
          'metadados': {'arquivo': str},
          'amostras': {
              '1/26': {'CI': [cq|None, ...], 'HPV16': [...], 'HPV18': [...], 'HPV_AR': [...]},
              ...
          },
          'controles': {
              'cn': {'CI': [...], 'HPV16': [...], 'HPV18': [...], 'HPV_AR': [...]},
              'cp': {'CI': [...], 'HPV16': [...], 'HPV18': [...], 'HPV_AR': [...]},
          },
          'por_poco': {posicao: {'CI': cq|None, ..., '_tipo': str, '_sample': str}},
        }

    Lança ``ValueError`` se o formato não for reconhecido.
    """
    import io
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f'Não foi possível abrir o arquivo XLSX: {exc}') from exc

    # Tenta a aba "0" primeiro (padrão do Amplio®), depois a primeira aba
    if '0' in wb.sheetnames:
        ws = wb['0']
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('Arquivo XLSX vazio.')

    # Detecta linha de cabeçalho procurando "Well" na linha
    header_idx = None
    col_idx: dict[str, int] = {}
    for i, row in enumerate(rows):
        if row and 'Well' in row and 'Target' in row and 'Sample' in row:
            header_idx = i
            col_idx = {str(h).strip(): j for j, h in enumerate(row) if h is not None}
            break

    if header_idx is None:
        raise ValueError(
            'Formato de XLSX não reconhecido. '
            'Esperado cabeçalho com colunas "Well, Fluor, Content, Target, Sample, '
            'Threshold Cycle( C(t) )".'
        )

    # Localiza a coluna de Ct (nome pode variar)
    ct_col = None
    for possible in ('Threshold Cycle( C(t) )', 'Threshold Cycle(C(t))', 'C(t)', 'Ct', 'Cq'):
        if possible in col_idx:
            ct_col = possible
            break
    if ct_col is None:
        # Tenta pelo índice 6 (padrão observado no arquivo de exemplo)
        ct_col = list(col_idx.keys())[6] if len(col_idx) > 6 else None
    if ct_col is None:
        raise ValueError('Coluna de Ct (Threshold Cycle) não encontrada no XLSX.')

    amostras: dict = {}
    controles: dict = {
        'cn': {c: [] for c in _CANAIS},
        'cp': {c: [] for c in _CANAIS},
    }
    por_poco: dict = {}

    for row in rows[header_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        if len(row) <= max(col_idx.values()):
            continue

        well    = row[col_idx['Well']]
        target  = row[col_idx['Target']]
        content = row[col_idx['Content']]
        sample  = row[col_idx['Sample']]
        cq_raw  = row[col_idx[ct_col]]

        if well is None or target is None:
            continue

        canal = resolver_canal(str(target).strip())
        if canal is None:
            continue

        cq   = _safe_cq(cq_raw)
        tipo = CONTENT_TYPE_MAP.get(str(content).strip() if content else '')
        if tipo is None:
            continue

        posicao = _normalizar_posicao(str(well))
        sample_str = str(sample).strip() if sample else ''

        if tipo == 'amostra':
            if sample_str not in amostras:
                amostras[sample_str] = {c: [] for c in _CANAIS}
            amostras[sample_str][canal].append(cq)
        elif tipo in ('cn', 'cp'):
            controles[tipo][canal].append(cq)

        if posicao not in por_poco:
            por_poco[posicao] = {c: None for c in _CANAIS}
            por_poco[posicao]['_tipo']   = tipo
            por_poco[posicao]['_sample'] = sample_str
        por_poco[posicao][canal] = cq

    return {
        'metadados': {'arquivo': filename},
        'amostras':  amostras,
        'controles': controles,
        'por_poco':  por_poco,
    }
