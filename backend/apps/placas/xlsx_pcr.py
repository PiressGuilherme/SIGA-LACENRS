"""
Geração do mapa de PCR preenchendo o template XLSX.

Preenche ``template_mapa_PCR.xlsx`` com dados da placa e retorna os
bytes do XLSX pronto para download.

Campos preenchidos:
  - D1: código da placa no título (substitui 'xxxxxx')
  - L4: código da placa no ensaio (substitui 'xxxxxx')
  - B4: kit de interpretação da placa
  - C6:N13: grid 8×12 com código interno das amostras / CP / CN
  - Linhas 17-21 (colunas B/D/E): nome do reagente, vol/reação, vol total
  - K17: Registro PCR = nome do usuário que gerou o mapa
  - K18: Operador PCR = em branco (preenchimento manual)

A tabela de reagentes usa o protocolo associado a cada grupo de reação da
placa. Se houver mais de um grupo, os volumes totais somam todos os grupos.
"""
import io
import os
from decimal import Decimal

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock

from .models import TipoConteudoPoco


def _substituir_celula(cell, old, new):
    """Substitui texto numa célula preservando rich text (runs com formatação)."""
    val = cell.value
    if isinstance(val, CellRichText):
        for run in val:
            if isinstance(run, TextBlock) and old in run.text:
                run.text = run.text.replace(old, new)
    elif isinstance(val, str):
        cell.value = val.replace(old, new)

ROWS        = list('ABCDEFGH')
COLS_PADDED = [f'{i:02d}' for i in range(1, 13)]

GRID_START_ROW = 6
GRID_START_COL = 3   # coluna C

# Linhas disponíveis para reagentes (B17:E21 — 5 linhas)
REAGENTES_START_ROW = 17
REAGENTES_MAX_ROWS  = 5

TIPO_LABELS = {
    TipoConteudoPoco.CONTROLE_NEGATIVO: 'CN',
    TipoConteudoPoco.CONTROLE_POSITIVO: 'CP',
}

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template_mapa_PCR.xlsx')


def _get_label(poco):
    label = TIPO_LABELS.get(poco.tipo_conteudo)
    if label is not None:
        return label
    if poco.amostra:
        return poco.amostra.codigo_interno
    return ''


def _calcular_reagentes(placa):
    """
    Retorna (n_reacoes, lista de (nome, vol_por_reacao, vol_total)).

    n_reacoes: total de poços não-vazios com margem (somado de todos os grupos).
    vol_total de cada reagente = vol_por_reacao × n_reacoes.
    """
    grupos = placa.grupo_reacoes.select_related('protocolo').prefetch_related(
        'protocolo__reagentes'
    ).all()

    totais: dict[str, Decimal] = {}
    vol_reacao: dict[str, Decimal] = {}
    n_reacoes_total = Decimal('0')

    for gr in grupos:
        proto = gr.protocolo
        margem = Decimal(str(proto.margem_percentual or 0)) / 100

        n = placa.pocos.filter(grupo=gr.grupo).exclude(
            tipo_conteudo=TipoConteudoPoco.VAZIO
        ).count()
        n_com_margem = Decimal(str(n)) * (1 + margem)
        n_reacoes_total += n_com_margem

        for reagente in proto.reagentes.order_by('ordem'):
            nome = reagente.nome
            vol  = Decimal(str(reagente.volume_por_reacao))
            totais[nome]     = totais.get(nome, Decimal('0')) + vol * n_com_margem
            vol_reacao[nome] = vol_reacao.get(nome, vol)

    reagentes = [(nome, vol_reacao[nome], totais[nome]) for nome in totais]
    return n_reacoes_total, reagentes


def gerar_xlsx_pcr(placa, operador=None):
    """Preenche o template PCR e retorna bytes do XLSX.

    Args:
        placa:    instância de Placa (tipo PCR)
        operador: instância de User — quem gerou o mapa ("Registro PCR")
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH, rich_text=True)
    ws = wb.active

    # Código da placa no título e ensaio (preserva rich text / formatação dos runs)
    _substituir_celula(ws['D1'], 'xxxxxx', placa.codigo)
    _substituir_celula(ws['L4'], 'xxxxxx', placa.codigo)

    # Kit de interpretação
    if placa.kit_interpretacao:
        ws['B4'].value = f'Kit: {placa.kit_interpretacao.nome}'

    # Grid 8×12
    pocos_map = {p.posicao: p for p in placa.pocos.select_related('amostra').all()}
    for ri, row_letter in enumerate(ROWS):
        for ci, col_num in enumerate(COLS_PADDED):
            pos = f'{row_letter}{col_num}'
            poco = pocos_map.get(pos)
            cell = ws.cell(row=GRID_START_ROW + ri, column=GRID_START_COL + ci)
            if poco and poco.tipo_conteudo != TipoConteudoPoco.VAZIO:
                cell.value = _get_label(poco)
            else:
                cell.value = None

    # Tabela de reagentes
    n_reacoes, reagentes = _calcular_reagentes(placa)

    # E16: número de reações (substitui "x reações")
    _substituir_celula(ws['E16'], 'x reações', str(round(float(n_reacoes), 1)))

    total_1rx = Decimal('0')
    total_geral = Decimal('0')
    for i, (nome, vol_rx, vol_total) in enumerate(reagentes[:REAGENTES_MAX_ROWS]):
        row = REAGENTES_START_ROW + i
        ws.cell(row=row, column=2).value = nome                        # B: reagente
        ws.cell(row=row, column=4).value = float(vol_rx)               # D: vol/reação
        ws.cell(row=row, column=5).value = round(float(vol_total), 1)  # E: vol total
        total_1rx   += vol_rx
        total_geral += vol_total

    # Linha TOTAL: D22 = soma de 1 reação, E22 = soma total
    ws['D22'].value = round(float(total_1rx), 1)
    ws['E22'].value = round(float(total_geral), 1)

    # Registro PCR = quem gerou o mapa
    registro_nome = ''
    if operador:
        registro_nome = getattr(operador, 'nome_completo', str(operador))
    ws['K17'].value = f'Registro PCR: {registro_nome}' if registro_nome else 'Registro PCR:'
    ws['K18'].value = 'Operador PCR:'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
